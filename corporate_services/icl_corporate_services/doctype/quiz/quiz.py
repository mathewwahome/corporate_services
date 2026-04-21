# Copyright (c) 2026, IntelliSOFT Consulting and contributors
# For license information, please see license.txt

import csv
import io
import os
import tempfile

import frappe
import openpyxl
from frappe.model.document import Document
from frappe.utils.file_manager import get_file


class Quiz(Document):
    pass


TEMPLATE_HEADERS = [
    "question",
    "question_type",
    "answer_options",
    "correct_answer",
    "disqualifying_answers",
    "explanation",
    "category",
]
TEMPLATE_ROWS = [
    [
        "What is the main purpose of the policy?",
        "Single Select",
        "Promote ethics|Replace contracts|Marketing guidance",
        "A",
        "",
        "Pick the best answer",
        "Policy",
    ],
    [
        "Which of the following apply?",
        "Multi Select",
        "Option one|Option two|Option three|Option four",
        "A|C",
        "",
        "Select all correct choices",
        "Policy",
    ],
    [
        "Describe the escalation process",
        "Open Ended",
        "",
        "",
        "",
        "Free text response",
        "Policy",
    ],
]


def _normalize_header(value):
    return (value or "").strip().lower().replace(" ", "_")


def _normalize_question_type(value):
    question_type = (value or "Single Select").strip()
    valid_types = {"Single Select", "Multi Select", "Open Ended"}
    if question_type not in valid_types:
        frappe.throw(f"Invalid question type: {question_type}")
    return question_type


def _split_pipe_values(value):
    if not (value or "").strip():
        return []
    return [item.strip() for item in str(value).replace("\n", "|").split("|") if item.strip()]


def _question_base_name(question_text):
    return (question_text or "")[:50].strip()


def _get_file_rows(file_url):
    file_doc = frappe.get_doc("File", {"file_url": file_url})
    file_content = get_file(file_doc.file_url)[1]
    extension = (file_doc.file_name or "").split(".")[-1].lower()

    if extension == "csv":
        csv_file = io.StringIO(file_content.decode("utf-8") if isinstance(file_content, bytes) else file_content)
        return list(csv.DictReader(csv_file))

    if extension in {"xls", "xlsx"}:
        with tempfile.NamedTemporaryFile(delete=False, suffix=f".{extension}") as temp_file:
            temp_file.write(file_content)
            temp_file_path = temp_file.name

        try:
            workbook = openpyxl.load_workbook(temp_file_path, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
        finally:
            os.unlink(temp_file_path)

        if not rows:
            return []

        headers = [_normalize_header(cell) for cell in rows[0]]
        parsed_rows = []
        for row in rows[1:]:
            if not row or not any(cell not in (None, "") for cell in row):
                continue
            parsed_rows.append(
                {
                    headers[index]: row[index]
                    for index in range(min(len(headers), len(row)))
                    if headers[index]
                }
            )
        return parsed_rows

    frappe.throw("Unsupported file format. Please upload a CSV or Excel file.")


def _build_question_doc(row):
    normalized_row = {_normalize_header(key): value for key, value in row.items()}
    question_text = (normalized_row.get("question") or "").strip()

    if not question_text:
        frappe.throw("Each imported row must include a question.")

    question_type = _normalize_question_type(normalized_row.get("question_type"))
    answer_options = "\n".join(_split_pipe_values(normalized_row.get("answer_options")))
    correct_answers = "\n".join(_split_pipe_values(normalized_row.get("correct_answer")))
    disqualifying_answers = "\n".join(_split_pipe_values(normalized_row.get("disqualifying_answers")))
    category = _ensure_category(normalized_row.get("category"))

    return {
        "question": question_text,
        "question_type": question_type,
        "answer_options": answer_options,
        "correct_answer": correct_answers,
        "disqualifying_answers": disqualifying_answers,
        "explanation": normalized_row.get("explanation"),
        "category": category,
    }


def _ensure_category(category_value):
    category_name = (category_value or "").strip()
    if not category_name:
        return None

    existing_name = frappe.db.get_value("Quiz Question Category", {"name1": category_name}, "name")
    if existing_name:
        return existing_name

    category = frappe.get_doc(
        {
            "doctype": "Quiz Question Category",
            "name1": category_name,
        }
    )
    category.insert()
    return category.name


def _get_existing_question(question_text):
    exact_match = frappe.db.get_value("Quiz Question", {"question": question_text}, "name")
    if exact_match:
        return frappe.get_doc("Quiz Question", exact_match)

    base_name = _question_base_name(question_text)
    if not base_name:
        return None

    candidates = frappe.get_all(
        "Quiz Question",
        filters={"name": ["like", f"{base_name}%"]},
        fields=["name", "question"],
        order_by="creation asc",
    )

    if not candidates:
        return None

    if len(candidates) == 1:
        return frappe.get_doc("Quiz Question", candidates[0].name)

    for candidate in candidates:
        if candidate.name == base_name:
            return frappe.get_doc("Quiz Question", candidate.name)

    return frappe.get_doc("Quiz Question", candidates[0].name)


def _question_has_changes(question_doc, question_data):
    comparable_fields = [
        "question",
        "question_type",
        "answer_options",
        "correct_answer",
        "disqualifying_answers",
        "explanation",
        "category",
    ]
    return any((getattr(question_doc, field, None) or "") != (question_data.get(field) or "") for field in comparable_fields)


def _upsert_question(row):
    question_data = _build_question_doc(row)
    existing_question = _get_existing_question(question_data["question"])

    if existing_question:
        changed = _question_has_changes(existing_question, question_data)
        if changed:
            existing_question.update(question_data)
            existing_question.save()
            return existing_question, "updated"
        return existing_question, "skipped"

    question_doc = frappe.get_doc({"doctype": "Quiz Question", **question_data})
    question_doc.insert()
    return question_doc, "created"


def _create_file(file_name, content):
    file_doc = frappe.get_doc(
        {
            "doctype": "File",
            "file_name": file_name,
            "content": content,
            "is_private": 0,
        }
    )
    file_doc.save(ignore_permissions=True)
    return file_doc.file_url


@frappe.whitelist()
def download_quiz_import_template(file_type="csv"):
    file_type = (file_type or "csv").lower()

    if file_type == "csv":
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(TEMPLATE_HEADERS)
        writer.writerows(TEMPLATE_ROWS)
        file_url = _create_file("quiz_question_import_template.csv", buffer.getvalue())
        return {"file_url": file_url}

    if file_type == "xlsx":
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Quiz Questions"
        sheet.append(TEMPLATE_HEADERS)
        for row in TEMPLATE_ROWS:
            sheet.append(row)

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        file_url = _create_file("quiz_question_import_template.xlsx", output.getvalue())
        return {"file_url": file_url}

    frappe.throw("Unsupported template type. Use csv or xlsx.")


@frappe.whitelist()
def import_questions_to_quiz(docname):
    quiz = frappe.get_doc("Quiz", docname)

    if not quiz.question_import_file:
        frappe.throw("Attach a CSV or Excel file before importing questions.")

    rows = _get_file_rows(quiz.question_import_file)
    if not rows:
        frappe.throw("No rows were found in the uploaded import file.")

    created_count = 0
    updated_count = 0
    skipped_count = 0
    linked_count = 0
    existing_questions = {row.question for row in quiz.questions}
    created_categories_before = set(frappe.get_all("Quiz Question Category", pluck="name"))

    for row in rows:
        question_doc, action = _upsert_question(row)

        if question_doc.name not in existing_questions:
            quiz.append(
                "questions",
                {
                    "question": question_doc.name,
                    "question_text": question_doc.question,
                },
            )
            existing_questions.add(question_doc.name)
            linked_count += 1

        if action == "created":
            created_count += 1
        elif action == "updated":
            updated_count += 1
        else:
            skipped_count += 1

    quiz.save()
    frappe.db.commit()

    created_categories_after = set(frappe.get_all("Quiz Question Category", pluck="name"))
    new_categories = sorted(created_categories_after - created_categories_before)

    return {
        "created_count": created_count,
        "updated_count": updated_count,
        "skipped_count": skipped_count,
        "linked_count": linked_count,
        "categories_created": new_categories,
        "quiz": quiz.name,
    }
