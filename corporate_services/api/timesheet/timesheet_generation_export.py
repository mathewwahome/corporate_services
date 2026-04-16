import frappe
import calendar
from datetime import datetime, timedelta
from io import BytesIO

from frappe.utils.file_manager import save_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


DEFAULT_TEMPLATE = "Default"
SHORT_TERM_CONSULTANT_TEMPLATE = "Short Term Consultant"


def get_employee_timesheet_template(employee):
    contract_type = frappe.db.get_value("Employee", employee, "custom_contract_type")
    if not contract_type:
        return DEFAULT_TEMPLATE

    return (
        frappe.db.get_value("Contract Type", contract_type, "timesheet_template")
        or DEFAULT_TEMPLATE
    )


def get_timesheet_period(month, year):
    if month == 1:
        previous_year = year - 1
        previous_month = 12
    else:
        previous_year = year
        previous_month = month - 1

    last_day = calendar.monthrange(previous_year, previous_month)[1]
    start_date = datetime(previous_year, previous_month, min(29, last_day)).date()
    end_date = datetime(year, month, 28).date()

    return start_date, end_date


def get_projects_for_user(user_id):
    projects_list = []
    if not user_id:
        return projects_list

    projects = frappe.get_all("Project", fields=["name", "project_name"])
    for project in projects:
        projects_user = frappe.get_all(
            "Project User",
            filters={"parent": project["name"], "user": user_id},
        )
        if projects_user:
            projects_list.append([project["project_name"]])
            projects_list.append([])
            projects_list.append([])

    return projects_list


def build_default_template_workbook(user_id, start_date, end_date):
    days = []
    dates = []
    current_date = start_date
    while current_date <= end_date:
        days.append(current_date.strftime("%A"))
        dates.append(current_date.strftime("%d"))
        current_date += timedelta(days=1)

    projects_list = get_projects_for_user(user_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"

    week = ["", ""] + days + ["Total Hours"]
    header = ["Projects", "Tasks"] + dates + ["Total Hours"]
    ws.append(week)
    ws.append(header)

    project_rows = []
    current_row = 3
    for row in projects_list:
        ws.append(row)
        if row and row[0]:
            project_rows.append(current_row)
        current_row += 1

    blue_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col_num in range(1, len(header) + 1):
        cell = ws.cell(row=2, column=col_num)
        cell.fill = blue_fill
        cell.border = thin_border

    for row_num in project_rows:
        for col_num in range(1, len(header) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = blue_fill
            cell.border = thin_border

    activity_types = frappe.get_all(
        "Activity Type", filters={"disabled": 0}, fields=["name"]
    )
    activity_types = [
        activity for activity in activity_types if activity["name"].lower() != "projects"
    ]

    activity_rows_start = len(projects_list) + 3
    activity_rows = []
    for i, activity in enumerate(activity_types):
        row_num = activity_rows_start + i * 2
        ws.cell(row=row_num, column=1, value=activity["name"])
        activity_rows.append(row_num)

    current_max_row = ws.max_row
    for i in range(3):
        empty_row_num = current_max_row + 1 + i
        for col_num in range(1, len(header) + 1):
            ws.cell(row=empty_row_num, column=col_num, value="")

    max_row = ws.max_row
    total_hours_col = len(header)
    start_col = 3
    end_col = len(dates) + 2

    for row_num in range(3, max_row + 1):
        start_col_letter = ws.cell(row=row_num, column=start_col).column_letter
        end_col_letter = ws.cell(row=row_num, column=end_col).column_letter
        formula = f"=SUM({start_col_letter}{row_num}:{end_col_letter}{row_num})"
        ws.cell(row=row_num, column=total_hours_col, value=formula)
        ws.cell(row=row_num, column=total_hours_col).border = thin_border

    for col_num, day in enumerate(days, start=3):
        if day in ("Saturday", "Sunday"):
            for row_num in range(1, max_row + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.fill = blue_fill
                cell.border = thin_border

    for row_num in activity_rows:
        for col_num in range(1, len(header) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = blue_fill
            cell.border = thin_border

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=len(header)):
        for cell in row:
            cell.border = thin_border

    total_row = max_row + 1
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=1).fill = blue_fill
    ws.cell(row=total_row, column=2).fill = blue_fill

    for col_num in range(3, len(dates) + 3):
        col_letter = ws.cell(row=1, column=col_num).column_letter
        formula = f"=SUM({col_letter}3:{col_letter}{max_row})"
        ws.cell(row=total_row, column=col_num, value=formula)
        ws.cell(row=total_row, column=col_num).fill = blue_fill
        ws.cell(row=total_row, column=col_num).border = thin_border

    start_col_letter = ws.cell(row=1, column=3).column_letter
    end_col_letter = ws.cell(row=1, column=len(dates) + 2).column_letter
    total_hours_total_formula = (
        f"=SUM({start_col_letter}{total_row}:{end_col_letter}{total_row})"
    )
    ws.cell(row=total_row, column=total_hours_col, value=total_hours_total_formula)
    ws.cell(row=total_row, column=total_hours_col).fill = blue_fill
    ws.cell(row=total_row, column=total_hours_col).border = thin_border

    return wb


def build_short_term_consultant_workbook(employee_name, month_name, year, start_date, end_date):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month_name[:3]} {year}"

    dark_blue_fill = PatternFill(start_color="FF004C7F", end_color="FF004C7F", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    title_font = Font(name="Helvetica Neue", size=12, bold=True)
    bold_font = Font(name="Helvetica Neue", bold=True)
    total_font = Font(name="Arial", bold=True)

    ws.column_dimensions["A"].width = 22.14
    ws.column_dimensions["B"].width = 53.43
    ws.column_dimensions["C"].width = 68.29
    ws.column_dimensions["D"].width = 11.71

    ws.merge_cells("A1:D1")
    ws.merge_cells("A2:D2")
    ws.merge_cells("A3:D3")

    ws["A1"] = "INTELLISOFT CONSULTING LTD"
    ws["A2"] = "Timesheet"
    ws["A3"] = "Time Sheet"
    ws["A4"] = "Consutant Name:"
    ws["B4"] = employee_name
    ws["A5"] = "Period of timesheet"
    ws["B5"] = "{} - {}".format(
        start_date.strftime("%d %b %Y"), end_date.strftime("%d %b %Y")
    )

    for cell_ref in ("A1", "A2"):
        ws[cell_ref].font = title_font
        ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center")
        ws[cell_ref].fill = white_fill

    for cell_ref in ("A1",):
        ws[cell_ref].border = thin_border

    ws["A3"].font = bold_font
    ws["A3"].fill = dark_blue_fill
    ws["A3"].font = Font(name="Helvetica Neue", bold=True, color="FFFFFFFF")

    for cell_ref in ("A4", "A5"):
        ws[cell_ref].font = bold_font

    headers = ["Date", "Task", "Deliverables", "Hours Worked"]
    for index, header in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=index, value=header)
        cell.fill = dark_blue_fill
        cell.font = Font(name="Helvetica Neue", bold=True, color="FFFFFFFF")
        cell.border = thin_border

    work_dates = []
    current_date = start_date
    while current_date <= end_date:
        if current_date.weekday() < 5:
            work_dates.append(current_date)
        current_date += timedelta(days=1)

    data_start_row = 8
    data_end_row = data_start_row + len(work_dates) - 1

    for offset, work_date in enumerate(work_dates):
        row_num = data_start_row + offset
        date_cell = ws.cell(row=row_num, column=1)
        date_cell.value = work_date
        date_cell.number_format = "mm/dd/yyyy"
        date_cell.alignment = Alignment(horizontal="center", vertical="top")
        date_cell.border = thin_border

        task_cell = ws.cell(row=row_num, column=2)
        task_cell.alignment = Alignment(horizontal="left")
        task_cell.border = thin_border

        deliverables_cell = ws.cell(row=row_num, column=3)
        deliverables_cell.alignment = Alignment(horizontal="left")
        deliverables_cell.border = thin_border

        hours_cell = ws.cell(row=row_num, column=4)
        hours_cell.alignment = Alignment(horizontal="right")
        hours_cell.border = thin_border

    total_row = data_end_row + 1
    total_label_cell = ws.cell(row=total_row, column=2, value="Total Hours worked")
    total_label_cell.font = total_font
    total_label_cell.alignment = Alignment(horizontal="right", vertical="top")

    total_hours_cell = ws.cell(
        row=total_row,
        column=4,
        value=f"=SUM(D{data_start_row}:D{data_end_row})",
    )
    total_hours_cell.alignment = Alignment(horizontal="right", vertical="top")

    return wb

@frappe.whitelist()
def timesheet_generation_export(docname):
    try:
        doc = frappe.get_doc("Timesheet Submission", docname)

        employee = doc.employee
        employee_name = frappe.db.get_value("Employee", employee, "employee_name")
        user_id = frappe.db.get_value("Employee", employee, "user_id")
        template_name = get_employee_timesheet_template(employee)

        date_str = doc.month_year
        month = int(date_str.split("-")[0])
        year = int(date_str.split("-")[1])
        month_name = datetime(year, month, 1).strftime('%B')
        start_date, end_date = get_timesheet_period(month, year)

        if template_name == SHORT_TERM_CONSULTANT_TEMPLATE:
            wb = build_short_term_consultant_workbook(
                employee_name=employee_name,
                month_name=month_name,
                year=year,
                start_date=start_date,
                end_date=end_date,
            )
        else:
            wb = build_default_template_workbook(
                user_id=user_id,
                start_date=start_date,
                end_date=end_date,
            )

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        file_name = f"{employee_name}-{month_name}{year}-Timesheet.xlsx"
        file_url = save_file(file_name, output.read(), "Timesheet Submission", docname, is_private=0)

        return file_url

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "timesheet_generation_export")
        return "error"
